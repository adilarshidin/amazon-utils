import pandas as pd
import glob
import os
from openpyxl import load_workbook
from pathlib import Path

# =========================
# PATHS
# =========================
CSV_FILE = "output/all_listings_ready.csv"
XLSX_DIR = "templates/worten"
OUTPUT_DIR = "output/worten_filled"

Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

# =========================
# AMAZON → WORTEN MAPPING
# =========================
WORTEN_MAPPING = {
    "moda": {
        "APPAREL", "ONE_PIECE_OUTFIT", "COAT", "SHIRT", "PANTS", "SHORTS", "SKIRT",
        "SWEATER", "SWEATSHIRT", "VEST", "SOCKS", "UNDERPANTS", "BASE_LAYER_APPAREL_SET",
        "HAT", "SCARF", "EARMUFF", "SWIMWEAR", "SNOWSUIT", "APPAREL_GLOVES", "SPORT_ACTIVITY_GLOVE",
        "BELTS", "SUSPENDER", "HANDBAG", "HIP_FLASK", "WATCH", "SHOE_ACCESSORY",
        "APPAREL_BELT", "APPAREL_HEAD_NECK_COVERING", "BRACELET", "SMOKING_PIPE"
    },

    "ropa_y_calzado_deportivo": {
        "SHOES", "BOOT", "SPORT_HELMET", "AUTOMOTIVE_HELMET", "SPORTING_GOODS",
        "SPORT_EQUIPMENT_BAG_CASE", "KNEE_PAD", "SAFETY_GLASSES", "SNOWSHOE",
        "SPORT_BAT", "SHOE_BAG", "CELL_PHONE_HOLSTER", "WAIST_PACK"
    },

    "salud_bienestar_y_cuidados_para_bebe": {
        "FIRST_AID_KIT", "SELF_DEFENSE_SPRAY", "PROTECTIVE_GLOVE",
        "SAFETY_HARNESS", "WASTE_BAG", "PET_TOY"
    },

    "productos_de_cuidado_personal": {
        "COSMETIC_CASE", "BODY_PAINT", "HEALTH_PERSONAL_CARE",
        "CUIDADO_BUCAL", "CUIDADO_CARA", "CUIDADO_CABELLO"
    },

    "supermercado_bebidas_y_limpieza": {
        "FOOD", "CLEANING_AGENT", "PEST_CONTROL_DEVICE", "SOLID_FIRE_FUEL"
    },

    "muebles_y_accesorios": {
        "HOME_FURNITURE_AND_DECOR", "STOOL_SEATING", "TABLE",
        "STORAGE_BOX", "CADDY", "NETTING_COVER"
    },

    "deporte_aire_libre_y_viaje": {
        "TENT", "TARP", "HAMMOCK", "SLEEPING_BAG", "SLEEPING_MAT",
        "OUTDOOR_RECREATION_PRODUCT", "BACKPACK", "DUFFEL_BAG", "CARRIER_BAG_CASE",
        "HYDRATION_PACK", "AIR_MATTRESS", "ANCHOR_STAKE", "BICYCLE_LIGHT",
        "CAMPING_EQUIPMENT", "CARGO_STRAP", "NAVIGATION_COMPASS",
        "PORTABLE_STOVE", "SURVIVAL_KIT", "MESS_KIT",
        "AIR_GUN_PROJECTILE", "GUN_CLEANING_KIT", "GUN_HOLSTER", "GUN_SLING",
        "BINOCULAR", "WEAPON_CASE"
    },

    "bricolaje_y_construccion": {
        "AXE", "KNIFE", "MULTITOOL", "SCREW_GUN", "SHOVEL_SPADE", "SAW",
        "ADHESIVE_TAPES", "BUCKET", "CABLE_TIE", "CORD_ROPE",
        "ELASTIC_BAND", "FABRIC_APPLIQUE_PATCH", "KNIFE_SHARPENER", "SEWING_TOOL_SET",
        "CARABINER", "LOCK", "PAINT"
    },

    "hogar": {
        "COOKING_POT", "KITCHEN_KNIFE", "FLATWARE", "DISHWARE_PLACE_SETTING", "DISHWARE_PLATE",
        "THERMOS", "DRINKING_CUP", "BOTTLE", "TOWEL", "PILLOW", "BLANKET", "STORAGE_BAG",
        "BLADED_FOOD_PEELER", "CAN_OPENER", "KITCHEN", "MANUAL_FOOD_MILL_GRINDER", "BUCKLE"
    },

    "merchandising_&_gifting": {
        "GLITTER", "CHARM", "BADGE_HOLDER", "BANNER", "KEYCHAIN", "LABEL"
    },

    "smart_home": {
        "LIGHT_BULB", "LIGHT_FIXTURE", "HOME_LIGHTING_AND_LAMPS",
        "FLASHLIGHT", "UTILITY_HOLSTER_POUCH"
    },

    "fotografia_y_video": {
        "ACCESSORIOS_FOTOGRAFIA_Y_VIDEO",
        "CAMERAS", "OBJETIVOS_Y_FLASHES", "VIDEO",
        "CAMERA_FILM"
    },

    "mascotas": {
        "ACCESORIOS_PARA_ANIMALES", "COMIDA_PARA_ANIMALES",
        "HIGIENE_CUIDADO_Y_SALUD_PARA_ANIMALES", "JUGUETES_PARA_ANIMALES",
        "LOCALIZADORES_Y_SEGURIDAD_DE_ANIMALES", "MUEBLES_PARA_ANIMALES",
        "TRANSPORTE_DE_ANIMALES"
    },

    "electrodomesticos": {
        "AIRE_ACONDICIONADO_Y_CALOR", "CAFETERAS", "COCINAS",
        "CONGELADORES", "EQUIPOS_INDUSTRIALES", "FRIGORIFICOS_Y_NEVERAS",
        "INTEGRABLES", "LAVADORAS", "LAVAVAJILLAS", "LIMPIEZA_DE_SUPERFICIES",
        "MICROONDAS_Y_MINI_HORNOS", "PEQUENOS_ELECTRODOMESTICOS", "TRATAMIENTO_DE_ROPA",
        "VINOTECAS", "SPACE_HEATER"
    },

    "equipamiento_y_piezas_de_vehiculos": {
        "MOTOR_ENGINE_FUEL_TANK"
    },

    "musica": {
        "BRASS_AND_WOODWIND_INSTRUMENTS"
    },

    "libros_y_audiolibros": {
        "BLANK_BOOK"
    }
}

WORTEN_CATEGORY_MAPPING = {
    "moda": {
        "Accesorios": [
            "Billeteros y Bolsas para Carteras",
            "Cinturones",
            "Corbatas y Lazos",
            "Gafas Bloqueo de Luz",
            "Gafas de Sol",
            "Llaveros",
            "Otros Accesorios",
            "Pañuelos y Bufandas",
            "Paraguas",
            "Sombreros, Gorras y Guantes"
        ],
        "Joyería - Acero, Metal, Latón, Otros": [
            "Anillos (acero, metal, latón)",
            "Collares (acero, metal, latón)",
            "Otras Piezas de Joyería",
            "Pendientes (acero, metal, latón)",
            "Pulseras (Acero, Metal, Latón)"
        ],
        "Ropa": [
            "Abrigos y sobretodos",
            "Bikinis y bañadores para la playa",
            "Blusas y camisas",
            "Camisetas, tops, sudaderas y jerséis",
            "Chaquetas y chalecos",
            "Conceptos básicos de la ropa",
            "Faldas y pantalones cortos",
            "Mallas de ropa",
            "Outra Ropa",
            "Pantalones y vaqueros",
            "Vestidos, vestidos y jerseys"
        ],
        "Ropa interior": [
            "Bodies",
            "Calcetines",
            "Calzoncillos/bóxers",
            "Camisetas interiores",
            "Conjuntos de bragas y sujetador",
            "Faldas",
            "Ligueros",
            "Medias",
            "Sujetadores"
        ],
        "Zapatos": [
            "Botas y Botines",
            "Otros Zapatos",
            "Sandalias",
            "Zapatillas casuales",
            "Zapatos"
        ]
    },
    "ropa_y_calzado_deportivo": {
        "Accesorios de Moda Deportiva": [
            "Accesorios para el Cuello",
            "Accesorios para la Cabeza",
            "Calcetines Deportivos",
            "Cinturones Deportivos",
            "Gafas Deportivas",
            "Guantes Deportivos",
            "Más Accesorios para Moda Deportiva"
        ],
        "Equipamiento Deportivo Oficial": [
            "Equipamiento de Clubes",
            "Equipamiento Deportivo Oficial",
            "Más Artículos de Merchandising Deportivo",
            "Merchandising deportivo"
        ],
        "Ropa deportiva": [
            "Accesorios para el Calzado",
            "Botas de Fútbol",
            "Botas Deportivas",
            "Chancletas y Sandalias",
            "Mantenimiento del Calzado",
            "Más Calzado Deportivo",
            "Zapatillas de Nieve",
            "Zapatillas Deportivas",
            "Zapatos Acuáticos"
        ],
        "Ropa Deportiva": [
            "Camisetas, Polos y Tops",
            "Chandals y Conjuntos",
            "Chaquetas, Cazadoras y Cortavientos",
            "Faldas y vestidos",
            "Más Ropa Deportiva",
            "Pantalones, Pantalones Cortos y Mallas",
            "Ropa Interior Deportiva",
            "Sudaderas, Camisetas y Sudaderas con Capucha",
            "Trajes de una pieza"
        ]
    },
    "salud_bienestar_y_cuidados_para_bebe": {
        "Ayuda para la Memoria y Monitores de diagnóstico": [
            "Accesorios para gestionar, almacenar, dividir y aplastar tabletas",
            "Outros Acessórios e Auxiliares de Memória, Medicação e Acompanhamento",
            "Sensores para Colchones, Sillas y Almohadas"
        ],
        "Ayuda para Movilidad": [
            "Andadores",
            "Ayudas para la movilidad (artículos regulados para personas con discapacidad)",
            "Bastones y Muletas",
            "Ejercitadores de piernas y brazos y cintas de rehabilitación",
            "Outros Acessórios e Auxiliares de Mobilidade",
            "Rampas de movilidad",
            "Silla de ruedas",
            "Sillas de ruedas eléctricas"
        ],
        "Ayudas, Cuidado y Movilidad de Mayores": [
            "Accesibilidad, Apoyo y Ayudas para la Manipulación (Artículos Regulados para Personas con Discapacidad o Personas con Discapacidad)",
            "Asientos de elevación",
            "Barras de Apoyo Seguridad",
            "Calzadores, cordones y otros accesorios",
            "Camas y somieres ortopédicos y reclinables",
            "Colchones antiescaras",
            "Otros accesorios y ayudas para la accesibilidad, el apoyo y la manipulación",
            "Pinzas para agarrar objetos",
            "Sillones y sillas ortopédicas y reclinables",
            "Taburetes y asientos de ducha y baño",
            "Tobilleras, rodilleras, férulas y cabestrillos"
        ],
        "Cuidado de la vista": [
            "Ayuda para escritura",
            "Gafas Graduadas",
            "Lentes de Contacto",
            "Lupas y Gafas de lectura",
            "Otros accesorios y ayudas visuales"
        ],
        "Cuidado del oído y Amplificadores": [
            "Amplificadores de audición",
            "Dispositivo de limpieza de oídos",
            "Limpiador de oídos",
            "Otros accesorios y audífonos",
            "Tapones para los oídos"
        ],
        "Equipos y Accesorios de Salud": [
            "Apósitos de gasa, algodón hidrófilo y adhesivos",
            "Báscula de análisis corporal",
            "Básculas de baño",
            "Electro-estimuladores",
            "Equipos para la terapia de luz y sonido",
            "Equipos y accesorios de salud (artículos reglamentados para personas con discapacidad o personas con discapacidad)",
            "Estetoscopios",
            "Gel desinfectante (alcohol> 70%)",
            "Geriatría",
            "Guantes, viseras y otros equipos de protección",
            "Humidificadores para bebés",
            "Limpieza y desinfección",
            "Linterna Medica",
            "Martillos de reflejos",
            "Masajeador Cuero Cabelludo",
            "Masajeadores",
            "Máscaras quirúrgicas",
            "Mascarillas de tela y no quirúrgicas",
            "Medias de descanso y mallas",
            "Medidores de Glucosa y Colesterol",
            "Monitor de sueño",
            "Monitores e Intercomunicadores",
            "Ortopedia",
            "Otro Equipo Medico",
            "Otros Accesorios Médicos",
            "Oxímetros",
            "Primeros auxilios",
            "Tensiómetro",
            "Termómetros para adultos",
            "Testes de Covid",
            "Tratamiento Capilar",
            "Tratamiento con infrarrojos",
            "Vehículos para discapacitados"
        ],
        "Especial Bebé": [
            "Accesorios Baño Bebés y Niños",
            "Accesorios de Alimentación para Bebé o Niños",
            "Accesorios para la lactancia",
            "Accesorios para sillas de coche y cochecitos",
            "Aerosol / Nebulizador",
            "Asientos para bebés o niños",
            "Bañeras y asientos para bebés",
            "Barreras, cerraduras y otros accesorios de seguridad infantil para el hogar",
            "Básculas para bebés",
            "Batidoras Alimentación Bebé",
            "Calienta Biberones",
            "Cambiadores de pañales",
            "Cambio de pañales e higiene del bebé",
            "Chupetes y portachupetes",
            "Cochecitos y triciclos",
            "Equipo de alimentación para bebés",
            "Esterilizador",
            "Hamacas, columpios y andadores para bebés",
            "Iluminación y lámparas de dormir Bebé",
            "Licuadoras Alimentación Bebé",
            "Marsupios y Hondas",
            "Nidos y Capazos",
            "Otros Accesorios Bebés y Niños",
            "Otros Equipos Bebés y Niños",
            "Pañales y toallitas para bebés",
            "Parques para bebés",
            "Robot de cocina Alimentación Bebé",
            "Sacaleches",
            "Sillas y Asientos de bicicleta para bebés y niños",
            "Sillas y Asientos de coche para bebés y niños",
            "Termómetros para bebés",
            "Textil Bebés y Niños",
            "Toallitas",
            "Vigilabebés"
        ],
        "Higiene": [
            "Apósitos para incontinencia",
            "Baño portatl",
            "Cepillo de Limpieza y esponjas",
            "Orinal",
            "Outros Acessórios e Auxiliares de Higiene",
            "Toallitas y tampones sanitarios"
        ],
        "Nutrición Infantil": [],
        "Productos para el bienestar y el alivio físico": [
            "Cremas para masajes musculares y articulares",
            "Repelentes de insectos"
        ],
        "Salud sexual": [
            "Condones / Métodos anticonceptivos",
            "Consoladores",
            "Juegos y juguetes eróticos / sexuales",
            "Lencería, Corsés, Ligas, Camisones",
            "Limpieza y Otros accesorios - Eróticos / Sexuales",
            "Lubricantes - Eróticos / Sexuales",
            "Perfumes, Aromas e Inciensos - Erótico/Sexual",
            "Tapones sexuales",
            "Velas, geles y aceites - Erótico/Sexual",
            "Vibradores / Estimuladores"
        ],
        "Vaporizadores y accesorios": [
            "Accesorios para vaporizadores",
            "Vaporizadores"
        ]
    },
    "productos_de_cuidado_personal": {
        "Cuidado bucal": [
            "Accesorios para equipos de higiene bucal",
            "Cepillos de dientes eléctricos",
            "Otros equipos Higiene Oral"
        ],
        "Cuidado de cara y cuerpo": [
            "Accesorios Equipos Cuidado de Cara y Cuerpo",
            "Accesorios para el cuidado facial y corporal",
            "Afeitadoras corporales",
            "Cortapelos y barberos",
            "Depiladoras",
            "Otros Acessorios de equipos para Afeitado y depilación"
        ],
        "Cuidado del cabello": [
            "Accesorios para secadores, planchas y rizadores de pelo",
            "Moldeadores y Rizadores",
            "Otros equipos para Cuidado del cabello",
            "Plancha de pelo",
            "Secadores y difusores de pelo"
        ]
    },
    "supermercado_bebidas_y_limpieza": {
        "Agua, Zumos y Refrescos": [
            "Aguas",
            "Refrescos",
            "Zumos y Néctares"
        ],
        "Cestas de alimentos": [
            "Cestas de alimentos"
        ],
        "Embutidos y Quesos": [
            "Charcutería",
            "Otras especialidades",
            "Quesos"
        ],
        "Lácteos y Bebidas Vegetales": [
            "Bebidas vegetales",
            "Mantequillas y cremas culinarias",
            "Productos lácteos"
        ],
        "Limpieza Del Hogar": [
            "Ambientadores e insecticidas",
            "Bolsas de basura",
            "Camillas y muelles",
            "Cubos y Bolsas De Basura",
            "Limpieza de Baños",
            "Limpieza de la cocina",
            "Limpieza de ropa y calzado",
            "Limpieza De Utensilios",
            "Limpieza General",
            "Papel higiénico y de cocina"
        ],
        "Tienda de alimentación": [
            "Aceite De Oliva, Aceite y Vinagre",
            "Alimentos Para Niños",
            "Aperitivos y Patatas Fritas",
            "Arroz, Pastas y Harinas",
            "Azúcar y Postres",
            "Café, Té y Cacao",
            "Cereales y Barritas",
            "Chocolates, Chicles y Caramelos",
            "Conservas, Patés y Productos Envasados",
            "Especias y condimentos",
            "Frutas secas y deshidratadas, aceitunas y altramuces",
            "Ingredientes veganos (soja, seitán, tofu)",
            "Miel, mermeladas y cremas",
            "Pastelería, panadería, galletas y bizcochos",
            "Sal",
            "Salsas y pulpas",
            "Sopas, Comidas y Preparaciones"
        ]
    },
    "muebles_y_accesorios": {
        "Cocina": [
            "Almacenaje Cocina",
            "Muebles de cocina, Islas, Carros de cocina",
            "Otros Accesorios Cocina"
        ],
        "Comedor": [
            "Accesorios para muebles de comedor de interior",
            "Juegos de muebles de comedor para interiores",
            "Mesas de comedor de interior",
            "Otros muebles de comedor de interior",
            "Sillas y bancos de interior"
        ],
        "Cuarto de Baño/WC": [
            "Accesorios de Muebles de baño y lavabo",
            "Muebles, Armarios de baño"
        ],
        "Despacho": [
            "Acessorios para Despacho y Oficina",
            "Mesas y escritorio",
            "Otros Muebles de Oficina",
            "Sillas de escritorio"
        ],
        "Dormitorio": [
            "Accesorios Almacenaje Dormitorio",
            "Cabeceros, Somieres y Estructuras",
            "Camas de dormitorio",
            "Colchones y Toppers",
            "Cómodas y Mesitas de Noche",
            "Otros Muebles de Dormitorio",
            "Packs de mobiliario para dormitorios (colchones/camas)",
            "Roperos y Armarios"
        ],
        "Muebles Bebé y Niño": [
            "Accesorios Muebles de Bebé y Niño",
            "Colchones para bebés y niños pequeños",
            "Cunas y camas para bebés y niños pequeños",
            "Otros Muebles de Bebé y Niño"
        ],
        "Muebles de Jardín": [
            "Accesorios de Muebles de jardín",
            "Cojines / colchones / puffs para exterior",
            "Columpios de exterior",
            "Cubiertas de protección para exteriores",
            "Decoración de jardín/exterior",
            "Guirnaldas, Lámparas, Bolardos, Luces Solares para Jardín/Exterior",
            "Juegos de muebles de comedor para exterior",
            "Mesas de exterior",
            "Sillas y bancos de exterior",
            "Sofás, Sofás de Paleta, sillones de exterior",
            "Tumbonas para Jardin"
        ],
        "Recibidor": [
            "Accesorios Muebles de Recibidor",
            "Bancos de Recibidor",
            "Otros Muebles de Recibidor",
            "Percheros",
            "Zapateros y Bancos zapateros"
        ],
        "Salón": [
            "Accesorios Muebles para TV",
            "Consolas, aparadores y estanterías",
            "Mesas de sala de estar",
            "Muebles para TV",
            "Otros Muebles de Salón",
            "Pufs, otomanas para Interior",
            "Sofás y sillones"
        ]
    },
    "deporte_aire_libre_y_viaje": {
        "Accesorios deportivos": [],
        "Artes marciales y deportes de combate": [
            "Accesorios para deportes de combate",
            "Equipamiento para artes marciales",
            "Recursos de formación sobre punzonado/punzonado"
        ],
        "Bolsas de Deporte, Bolsos y Mochilas": [],
        "Botellas y Termos": [],
        "Camping": [
            "Colchonetas de Camping",
            "Higiene en el camping",
            "Iluminación de camping",
            "Mobiliario de camping",
            "Mochilas y Otros Materiales de Camping",
            "Neveras portátiles",
            "Parrillas y Hornos de Camping",
            "Protección contra insectos y primeros auxilios",
            "Sacos de dormir",
            "Tiendas y Refugios",
            "Utensilios de cocina para acampar"
        ],
        "Ciclismo": [
            "Bicicletas (no eléctricas)",
            "Equipo de protección para bicicletas/ciclismo",
            "Piezas de bicicleta - Cadenas",
            "Piezas de bicicleta - Iluminación",
            "Piezas de bicicleta - Pedales",
            "Piezas de bicicleta - Timbre",
            "Piezas para bicicletas - Frenos",
            "Piezas para bicicletas - Neumáticos/ruedas",
            "Piezas para bicicletas - Puños",
            "Piezas para bicicletas - Sillines y accesorios para sillines",
            "Piezas para bicicletas - Sistema de cambios",
            "Rodillos de entrenamiento y Otros Accesorios para Bicicletas/Ciclismo",
            "Transporte de bicicletas - Remolques y asientos para niños"
        ],
        "Deportes acuáticos Natación/Surf/Buceo/SUP": [
            "Accesorios para deportes acuáticos",
            "Equipos de buceo",
            "Equipos de natación",
            "Tablas/esquís",
            "Trajes de baño/trajes de buceo/surf/bodyboard"
        ],
        "Deportes de caza": [
            "Accesorios de caza",
            "Cebos de caza / Chimeneas / Bengalas"
        ],
        "Deportes de equipo/Pista/Campo": [
            "Accesorios para deportes de pista",
            "Equipamiento deportivo para correr",
            "Fundas/Bolsas para equipos deportivos",
            "Pelotas deportivas",
            "Recinto deportivo/postes de portería"
        ],
        "Deportes de nieve": [
            "Accesorios para deportes de nieve",
            "Esquís y tablas de snowboard",
            "Trineos"
        ],
        "Deportes de raqueta": [
            "Accesorios para deportes de raqueta",
            "Accesorios/Piezas de recambio - Raquetas deportivas",
            "Raquetas deportivas"
        ],
        "Deportes de Tiro": [
            "Arcos deportivos con objetivo",
            "Dardos para deportes de tiro al blanco",
            "Equipamiento deportivo Target - Otros",
            "Objetivos deportivos"
        ],
        "Electronic Equipment for Sports and Outdoors": [],
        "Electrónica para Fitness": [
            "Accesorios de Electrónica Fitness",
            "Auriculares y Audífonos Deportivos",
            "Lectores Mp3 Deportivos",
            "Otros Equipos de Electrónica Fitness",
            "Pulseras de Actividad",
            "Relojes Entrenamiento y Deportivos"
        ],
        "Equipos de Musculación y Fitness": [
            "Bicicletas de spinning",
            "Bicicletas estáticas",
            "Cintas de correr para fitness",
            "Elípticas",
            "Equipamiento de musculación",
            "Máquinas de remo",
            "Material de Fitness",
            "Pesas Libres/Halteretas y Barras",
            "Pistolas de masaje",
            "Plataformas vibratorias",
            "Steps y Steppers"
        ],
        "Equipos de protección deportiva": [
            "Máscaras/gafas deportivas",
            "Protección acolchada para el cuerpo durante la práctica de deportes",
            "Protectores bucales deportivos"
        ],
        "Escalada, alpinismo y Trekking": [
            "Accesorios de escalada y alpinismo",
            "Material deportivo para trekking (senderismo)/escalada"
        ],
        "GPS y sistemas de navegación para deportes y actividades al aire libre": [],
        "Otros deportes": [
            "Accesorios para otros deportes",
            "Equipamiento para otros deportes"
        ],
        "Pesca artesanal/deportiva": [
            "Accesorios de pesca",
            "Anzuelos de pesca",
            "Boyas de pesca",
            "Cañas y conjuntos de pesca",
            "Cebos de pesca",
            "Sedal/hilo de pesca"
        ],
        "Playa": [
            "Carros de transporte para la playa",
            "Paravientos, carpas y toldos de playa",
            "Sillas de playa",
            "Sombrillas de playa",
            "Soportes para equipos de playa y otros accesorios",
            "Toallas de playa"
        ],
        "Skates/patinetes (no eléctricos)": [
            "Accesorios para skates/patinete",
            "Patines",
            "Skates/patinetes (no eléctricos)"
        ],
        "Taco de polo/Taco de billar/Taco de golf/Taco de hockey/Taco de béisbol": [
            "Tacos de golf/Tacos de hockey/Tacos de béisbol",
            "Tacos de golf/Tacos de hockey/Tacos de béisbol - Otros"
        ],
        "Vehículos": [
            "Drones profesionales"
        ],
        "Viaje": [
            "Adaptadores de viaje",
            "Bolsas y Maletas",
            "Candados de viaje",
            "Mochilas y Macutos",
            "Organizadores para equipaje",
            "Otros Equipos y Accesorios de Viaje"
        ],
        "Yoga/Pilates/Gimnasia": [
            "Colchonetas de yoga/gimnasio",
            "Equipos de gimnasia",
            "Otros equipos de gimnasio"
        ]
    },
    "bricolaje_y_construccion": {
        "Construcción y Madera": [
            "Carros, Escaleras de Mano, Rejillas y Andamios",
            "Fontanería y Evacuación de Aguas",
            "Fregaderos",
            "Grifos y Duchas",
            "Herrajes",
            "Lavamanos",
            "Maletas de Herramientas, Bancos de Trabajo y Almacenamiento en el Garaje",
            "Morteros, Yesos, Cementos, Ladrillos y Áridos",
            "Otros Equipos de Señalización y Seguridad para Bricolaje y Construcción",
            "Otros Equipos y Accesorios de Construcción y Carpintería",
            "Puertas y Sistemas de Apertura",
            "Ropa y Calzado de Bricolaje y Construcción",
            "Siliconas, Adhesivos, Aislantes e Impermeabilizantes",
            "Tejas, Tejados y Marquesinas",
            "Ventanas y tragaluces"
        ],
        "Electricidad": [
            "Accesorios para Powerstations",
            "Alargadores, Enchufes y Adaptadores",
            "Baterías y Cargadores",
            "Cuadros eléctricos y componentes",
            "Dispositivos de Medición y Detectores",
            "Energías renovables",
            "Faroles, Linternas de obra",
            "Hojas de electricidad",
            "Interruptores, Tomas de Corriente y Otros Accesorios de Iluminación",
            "Otros Equipos Eléctricos de Electricidad y Energía",
            "Otros Equipos No Eléctricos y Accesorios de Electricidad y Energía",
            "Pilas",
            "Powerstations",
            "Rieles y Tubos Eléctricos",
            "Sistemas de Alimentación Ininterrumpida"
        ],
        "Equipos de transporte/elevación/escalada": [
            "Carretillas - Sin motor",
            "Carretillas elevadoras",
            "Escaleras y Escaleras",
            "Plataformas/Andamios",
            "Transpaleta"
        ],
        "Herramientas y equipos industriales": [
            "Accesorios Herramientas Eléctricas",
            "Accesorios Herramientas Manuales",
            "Amoladoras angulares",
            "Aspiradoras industriales",
            "Consumibles Herramientas Eléctricas",
            "Consumibles Herramientas Manuales",
            "Cuchillos, cortadores y otras herramientas de corte X-Ato para bricolaje/construcción",
            "Destornilladores/Tornillos",
            "Espátulas",
            "Generadores",
            "Hormigoneras",
            "Lijadoras y cepilladoras eléctricas",
            "Limas y escofina",
            "Llaves",
            "Máquinas y pistolas eléctricas para pintar",
            "Martillos",
            "Martillos perforadores y demoledores",
            "Multiherramientas",
            "Navajas de bolsillo e Afeitar, X-actos Profesionales",
            "Otras herramientas eléctricas",
            "Otras herramientas manuales",
            "Pinzas/Alicates y Turcas",
            "Sierras y sierras de mano",
            "Sierras, Sierras Eléctricas y Caladoras",
            "Taladros/destornilladores"
        ],
        "Pinturas y Productos de mantenimiento": [
            "Bandejas, rodillos, cepillos y brochas",
            "Diluyentes, Limpieza y Droguería",
            "Otros Equipos y Accesorios de Pintura y Droguería",
            "Pinturas, imprimaciones, Barnices, Tratamiento y Mantenimiento"
        ],
        "Sanitarios": [
            "Bañeras",
            "Baños",
            "Bidés",
            "Duchas",
            "Grifos y Duchas",
            "Lavabos",
            "Lavavajillas"
        ],
        "Suelos, Baldosas y Revestimientos": [
            "Aislamiento, Molduras, Esquinas, Frisos, Perfiles y Zócalos",
            "Productos de Mantenimiento y Limpieza de Suelos",
            "Suelos y Revestimientos Exteriores",
            "Suelos y Revestimientos Interiores"
        ],
        "Taller/Garaje Almacenaje y Accesorios": [
            "Accesorios para mesas y bancos de trabajo",
            "Bancos de trabajo",
            "Caballetes",
            "Cajas/bolsas de herramientas",
            "Carros de herramientas",
            "Correas y fundas para herramientas",
            "Gabinetes y estanterías para taller/garaje",
            "Mesas de corte",
            "Mochilas y bolsas de herramientas"
        ]
    },
    "hogar": {
        "Cocina y Mesa": [
            "Accesorios para bebidas no alcohólicas",
            "Bandejas, ollas, sartenes y accesorios",
            "Bolsas y carros de la compra",
            "Cubiertos de mesa y de mesa",
            "Cuchillos de Cocina and Bloques de cuchillos",
            "Dispensadores de productos de limpieza",
            "Fiambreras, termos y cajas de almuerzo",
            "Hervidores y teteras",
            "Organización y conservación de la cocina",
            "Tazas, tazones, tazas y botellas",
            "Utensilios de cocina",
            "Utensilios de panadería",
            "Vajillas, platos y cuencos"
        ],
        "Decoración": [
            "Árboles de Navidad Artificiales",
            "Cajas y cestas decorativas",
            "Coronas y guirnaldas navideñas",
            "Decoraciones para paredes",
            "Difusores de aroma, aceites e incienso",
            "Espejos para el hogar",
            "Ganchos y percheros de pared",
            "Huchas",
            "Jarrones y jarras para interiores",
            "Marcos de fotos y álbumes",
            "Marcos, lienzos y papel pintado",
            "Otros Accesorios de Decoración",
            "Otros adornos y decoraciones navideñas",
            "Pantallas",
            "Piezas decorativas de interior para el hogar",
            "Plantas artificiales y flores artificiales",
            "Portavelas, linternas decorativas y accesorios para velas",
            "Relojes para el hogar",
            "Velas"
        ],
        "Iluminación de Interior": [
            "Focos LED/Paneles LED/Luces empotradas",
            "Iluminación Decorativa",
            "Lámparas",
            "Lámparas para el hogar",
            "Luces navideñas",
            "Otros Equipos de Iluminación",
            "Pantallas, Bases para Lámparas y Cables"
        ],
        "Textiles": [
            "Alfombras para el hogar (interior y exterior)",
            "Almohadas",
            "Barras de Cortinas",
            "Cojines Decorativos",
            "Cortinas",
            "Edredones",
            "Fundas de sofá, de sillón y colchones",
            "Mantas y colchas",
            "Otros textiles para el hogar",
            "Persianas para el hogar",
            "Ropa de cama",
            "Textiles de baño",
            "Textiles de mesa y cojines para silla"
        ]
    },
    "merchandising_&_gifting": {
        "Gifting": [
            "Gadgets",
            "Otros artículos de Regalo",
            "Regalos de oficina",
            "Regalos para el hogar"
        ],
        "Merchandising": [
            "Cartas (Merchandising)",
            "Funkos y figuras coleccionables",
            "Otros artículos de Merchandising",
            "Papelería (Merchandising)",
            "Textil (Merchandising)"
        ]
    },
    "smart_home": [
        "Accesorios de Equipos de Smart Home",
        "Cámaras y Sistemas de Vigilancia",
        "Electricidad inteligente",
        "Iluminación Inteligente",
        "Otros Equipos de Smart Home",
        "Sistema de Alarma, Sensores y Detectores"
    ],
    "fotografia_y_video": {
        "Accesorios Fotografía y Vídeo": [
            "Accesorios para trípodes",
            "Adaptadores y convertidores para fotografía",
            "Baterías de cámara",
            "Caja de luz",
            "Cámaras Acción",
            "Cargadores de cámara",
            "Correas para cámara",
            "Cubiertas y parasoles",
            "Filtros de fotografía",
            "Fundas y Mochilas",
            "Otros Accesorios de Fotografía y Vídeo",
            "Prismáticos",
            "Telescopios y microscopios",
            "Trípodes y Monópodes"
        ],
        "Cámaras": [
            "Cámara con lentes intercambiables",
            "Cámara Instantánea",
            "Cámaras Analógicas",
            "Cámaras Bridge y Evil",
            "Cámaras Compactas",
            "Camaras Desechables",
            "Cámaras Reflex",
            "Otras Cámaras"
        ],
        "Objetivos y Flashes": [
            "Flashes",
            "Objetivos"
        ],
        "Video": [
            "Cámaras 360",
            "Cámaras de Acción",
            "Otros Equipos de Vídeo",
            "Videocámaras"
        ]
    },
    "mascotas": {
        "Accesorios para animales": [
            "Collares de entrenamiento y otros accesorios de entrenamiento",
            "Otros accesorios para animales",
            "Ropa para mascotas"
        ],
        "Comida para Animales": [
            "Comederos y bebederos",
            "Comida y snacks",
            "Otros accesorios de alimentación",
            "Piensos para bovinos y aves de corral (industria)",
            "Suplementos alimenticios para animales"
        ],
        "Higiene, cuidado y salud para animales": [
            "Cajas de Arena para Animales",
            "Desparasitantes y otros medicamentos veterinarios",
            "Equipos de higiene y salud para animales",
            "Inodoro para animales y arena",
            "Otros accesorios para el cuidado e higiene de mascotas"
        ],
        "Juguetes para animales": [],
        "Localizadores y Seguridad de Animales": [
            "Equipos de seguimiento y seguridad para animales",
            "Otros accesorios de seguimiento y seguridad para mascotas"
        ],
        "Muebles para animales": [
            "Acuarios y terrarios",
            "Camas, perreras y mantas para animales",
            "Equipos de muebles para mascotas",
            "Gallineros",
            "Jaulas",
            "Otros Accesorios Muebles y Decoración para Animales",
            "Redes, vallas, puertas, rampas y parques de animales"
        ],
        "Transporte de Animales": [
            "Accesorios para el transporte de mascotas",
            "Correas, collares, arneses y bozales",
            "Equipo de transporte de animales"
        ]
    },
    "electrodomesticos": {
        "Aire Acondicionado y Calefacción": [
            "Aficionados locales",
            "Aire Acondicionado",
            "Aire Acondicionado Portátil",
            "Bombas de Calor",
            "Calderas",
            "Calefacción de biomasa < 50KW",
            "Calefacción y Accesorios Climatización",
            "Calefactores",
            "Calentadores de Gas",
            "Chimeneas y estufas",
            "Deshumidificadores, humidificadores y purificadores de aire",
            "Estufas de Gas",
            "Otros Equipos de Calentadores de Agua",
            "Otros Equipos de Tratamiento de Aire",
            "Radiadores de Aceite",
            "Termo eléctrico",
            "Ventiladores de Torre"
        ],
        "Cafeteras": [
            "Accesorios de Cafeteras",
            "Cafetera Automática",
            "Cafetera de Cápsulas",
            "Cafeteras (Gama Profesional)",
            "Cafeteras de Goteo",
            "Cafeteras Manuales",
            "Molinillo de Café",
            "Otras Cafeteras"
        ],
        "Cocinas": [
            "Accesorios de Cocinas",
            "Cocinas a Gas",
            "Cocinas Eléctricas",
            "Cocinas Mixtas",
            "Cocinas Portátiles",
            "Cocinas Semiprofesionales",
            "Otras Cocinas"
        ],
        "Congeladores": [
            "Accesorios de Congeladores",
            "Arcón Congelador"
        ],
        "Equipos Industriales": [
            "Cocinas Industriales",
            "Hornos Industriales",
            "Industrial Cold Equipment",
            "Lavavajillas Industriales",
            "Máquinas Industriales de Ropa",
            "Placas de Cocina Industriales",
            "Vitrinas para cocinas industriales"
        ],
        "Frigoríficos y Neveras": [
            "Accesorios de Frigoríficos",
            "Frigorífico Americano",
            "Frigoríficos Combi",
            "Frigoríficos con Congelador",
            "Frigoríficos Sin Congelador",
            "Otros Frigoríficos"
        ],
        "Integrables": [
            "Accesorios para Electrodomésticos Integrables",
            "Arcón Congelador Integrables",
            "Campanas Extractoras Integrables",
            "Extractores Integrables",
            "Frigoríficos Americanos Integrables",
            "Frigoríficos Combi Integrables",
            "Frigoríficos con Congelador Integrables",
            "Frigoríficos Sin Congelador Integrables",
            "Hornos Integrables",
            "Lavadoras Integrables",
            "Lavavajillas Integrables",
            "Microondas Integrables",
            "Otros Electrodomésticos Integrables",
            "Placas y Vitrocerámicas Integrables",
            "Vinotecas Integrables"
        ],
        "Lavadoras": [
            "Accesorios de Lavadoras",
            "Lavadora Secadora",
            "Lavadoras",
            "Otras Lavadoras y Secadoras",
            "Secadoras"
        ],
        "Lavavajillas": [
            "Accesorios de Lavavajillas",
            "Lavavajillas"
        ],
        "Limpieza de Superficies": [
            "Accesorios para aspiradoras y otros equipos de limpieza",
            "Aspirador con Bolsa",
            "Aspirador de Mano",
            "Aspirador Escoba",
            "Aspirador sin Bolsa",
            "Aspiradora de Água",
            "Limpiador de Vapor",
            "Limpiaventanas",
            "Otros Equipos Limpieza de Superficies",
            "Pistola a Vapor",
            "Robot Aspirador"
        ],
        "Microondas y Mini Hornos": [
            "Accesorios Microondas",
            "Microondas con Grill",
            "Mini Hornos",
            "Otros Microondas y Mini Hornos"
        ],
        "Pequeños Electrodomésticos": [
            "Acessorios de Pequeños Electrodomésticos",
            "Agitadores de leche",
            "Amasadoras",
            "Básculas de Cocina",
            "Batidoras de mano",
            "Batidoras de vaso",
            "Creperas",
            "Desayuno (Gama Profesional)",
            "Exprimidores",
            "Fondues",
            "Freidoras",
            "Grills y Planchas de Cocina",
            "Hervidores de agua",
            "Jarra de Agua y Purificadores",
            "Licuadoras",
            "Máquina de Gofres",
            "Máquina de Helado",
            "Máquina de Palomitas",
            "Otros Pequeños Electrodomésticos",
            "Panificadoras",
            "Picadoras",
            "Preparación de alimentos (Gama Profesional)",
            "Robots de Cocina",
            "Sandwicheras",
            "Tostadoras",
            "Yogurteras"
        ],
        "Tratamiento de Ropa": [
            "Accesorios para planchas y máquinas de planchar",
            "Centro de planchado",
            "Máquinas de Coser",
            "Otros Equipos Tratamiento de Ropa",
            "Planchas",
            "Prensas de Vapor",
            "Quitapelusas",
            "Tablas de Planchar"
        ],
        "Vinotecas": [
            "Accesorios de Vinotecas",
            "Vinotecas"
        ]
    },
    "equipamiento_y_piezas_de_vehiculos": {
        "Accesorios, Piezas y Otros Equipos para Vehículos": [
            "Aceites, otros líquidos y recipientes para automóviles",
            "Aire acondicionado y ventilación del automóvil",
            "Alfombras",
            "Ambientadores y decoraciones para el coche",
            "Antenas de coche",
            "Arrancadores y baterías de coches",
            "Asientos de coche",
            "Bombas de agua y circuitos de refrigeración para automóviles",
            "Bombas de aire para automóviles",
            "Cadenas de coche",
            "Cofres de techo / Barras / Portaequipajes para coche",
            "Deflectores de viento para automóviles",
            "Embrague y transmisión para automóviles",
            "Enfriadores y ventiladores para automóviles",
            "Equipos y Material de Seguridad, Prevención y Emergencia para Automóviles",
            "Espejos de coche",
            "Filtros para coches",
            "Frenos / Discos / Pastillas para Coches",
            "Fundas para coches",
            "Limpiaparabrisas y escobillas de limpiaparabrisas para automóviles",
            "Llantas y tapacubos de coche",
            "Luces del coche",
            "Manómetros para coches",
            "Mantenimiento y reparación de automóviles",
            "Neumáticos de coche",
            "Otros Accesorios y Repuestos",
            "Piezas de escape de automóvil",
            "Piezas y accesorios del compartimento del motor del automóvil",
            "Pintura de coches",
            "Productos de limpieza y mantenimiento de automóviles",
            "Suspensión y dirección del automóvil"
        ],
        "Accesorios, Repuestos y Otros Equipos para Motos": [
            "Alarmas y dispositivos antirrobo para motocicletas",
            "Baterías para motos",
            "Cascos para Motos",
            "Equipamiento y Protecciones para Motociclismo",
            "Fundas, Cubiertas y Alfombrillas para Motos",
            "Intercomunicadores para motocicletas",
            "Maletas para motos",
            "Otros Accesorios y Piezas para Motos",
            "Ropa De Moto"
        ],
        "Equipos Multimedia": [
            "Accesorios de Multimedia",
            "Altavoces Automóvil",
            "Amplificadores de coche",
            "Asistentes de aparcamiento",
            "Autoradios",
            "Cámaras delanteras, traseras y para bebés",
            "Detectores de cámaras y radares",
            "GPS / Sistemas de Navegación",
            "Otros Equipos Multimedia",
            "Subwoofers para coche",
            "Transmisores FM"
        ],
        "Vehículos especiales": [
            "Piezas de barcos"
        ]
    },
    "musica": {
        "CD, DVD, Vinilo y Blu-ray": [],
        "Instrumentos musicales": [
            "Amplificadores para instrumentos musicales",
            "Cuerdas para instrumentos musicales",
            "Instrumentos de arco",
            "Instrumentos de cuerda",
            "Instrumentos de percusión",
            "Instrumentos de viento",
            "Micrófonos",
            "Otros accesorios para instrumentos musicales",
            "Otros Instrumentos",
            "Partituras",
            "Pianos y teclados"
        ]
    },
    "libros_y_audiolibros": {
        "Accesorios de lectura": [],
        "Audiolibros": [],
        "Libros": [
            "Cómics y manga",
            "Libros antiguos y raros",
            "Libros de crimen, suspense, terror y fantasía",
            "Libros de humor",
            "Libros de no ficción (biografías y memorias + ensayos y crónicas)",
            "Libros de otros géneros literarios y de ficción",
            "Libros de poesía, cuentos y teatro",
            "Libros de romance y literatura contemporánea",
            "Libros encuadernados en pieles, seda u otras telas",
            "Libros para niños y jóvenes adultos",
            "Libros prácticos",
            "Libros técnicos",
            "Manuales y apoyo escolar"
        ]
    }
}

# =========================
# LOAD CSV
# =========================
df = pd.read_csv(CSV_FILE, dtype=str)

required_cols = {"seller-sku"}
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"Missing required CSV columns: {missing}")

# Keep track of unmatched SKUs
all_skus = set(df["seller-sku"].dropna())
matched_skus_total = set()

# =========================
# PROCESS EACH XLSX
# =========================
xlsx_files = glob.glob(os.path.join(XLSX_DIR, "*.xlsx"))

total_skus_written = 0
# =========================
# WRITE SKUS + ADDITIONAL COLUMNS
# =========================

# Columns mapping: CSV column → XLSX columns
COLUMN_MAPPING = {
    "item-name": ["product_name_pt_PT", "product_name_es_ES", "product_description_pt_PT",
                  "product_description_es_ES"],
    "seller-sku": ["ean"],
    "amazon_product_type_es": ["type_pt_PT", "type_es_ES"],
    "manufacturer": ["product-brand"],
    "mp_category": ["mp_category"]
}

# Max number of images to write
MAX_IMAGES = 12

for xlsx_path in xlsx_files:
    filename = os.path.splitext(os.path.basename(xlsx_path))[0]

    if filename not in WORTEN_MAPPING:
        print(f"⏭️ Skipping {filename}.xlsx (no mapping)")
        continue

    amazon_types = WORTEN_MAPPING[filename]

    matched_df = df[df["amazon_product_type"].isin(amazon_types)].copy()
    if matched_df.empty:
        print(f"⚠️ No matches for {filename}.xlsx")
        continue

    matched_skus_total.update(matched_df["seller-sku"].dropna().tolist())
    total_skus_written += len(matched_df)
    print(f"📦 Writing {len(matched_df)} SKUs → {filename}.xlsx")

    wb = load_workbook(xlsx_path)
    if "Data" not in wb.sheetnames:
        print(f"❌ Sheet 'Data' not found in {filename}.xlsx")
        continue

    ws = wb["Data"]

    # Build column index mapping from header row 2
    col_index = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=2, column=col).value
        if header:
            col_index[header] = col

    # Ensure required XLSX columns exist
    for csv_col, xlsx_cols in COLUMN_MAPPING.items():
        for xlsx_col in xlsx_cols:
            if xlsx_col not in col_index:
                raise ValueError(f"'{xlsx_col}' column not found in {filename}.xlsx")

    # Find first empty row (starting from row 3)
    row = 3
    while ws.cell(row=row, column=col_index["product_id"]).value:
        row += 1

    # Write data row by row
    for _, row_data in matched_df.iterrows():
        # product_id (seller-sku)
        ws.cell(row=row, column=col_index["product_id"], value=row_data["seller-sku"])

        # CSV → XLSX columns mapping
        for csv_col, xlsx_cols in COLUMN_MAPPING.items():
            # Skip mp_category here since we'll handle it separately
            if csv_col == "mp_category":
                continue
            for xlsx_col in xlsx_cols:
                ws.cell(row=row, column=col_index[xlsx_col], value=row_data.get(csv_col, ""))

        # Images (image1..image12)
        for i in range(1, MAX_IMAGES + 1):
            csv_image_col = f"image{i}"
            xlsx_image_col = f"image{i}"
            if csv_image_col in row_data and xlsx_image_col in col_index:
                ws.cell(row=row, column=col_index[xlsx_image_col], value=row_data[csv_image_col])

        # mp_category value based on XLSX file
        category_value = WORTEN_CATEGORY_MAPPING.get(filename, "")
        ws.cell(row=row, column=col_index["mp_category"], value=category_value)

        row += 1

    # Save to output directory
    output_path = os.path.join(OUTPUT_DIR, os.path.basename(xlsx_path))
    wb.save(output_path)

# =========================
# REPORT UNMATCHED SKUS
# =========================
unmatched_skus = all_skus - matched_skus_total
if unmatched_skus:
    print(f"⚠️ {len(unmatched_skus)} SKUs were not processed:")
    for sku in sorted(unmatched_skus):
        print(f" - {sku}")

print(f"✅ All applicable Worten sheets updated successfully in {OUTPUT_DIR}.")
print(f"Total SKUs written: {total_skus_written}")

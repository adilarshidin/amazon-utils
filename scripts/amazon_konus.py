import re
import logging
import pandas as pd
import json
import os
from dotenv import load_dotenv
from mistralai import Mistral
from openpyxl import load_workbook
from pathlib import Path

# ---------------- CONFIG ----------------
excel_path = "templates/konus.xlsm"
sheet_name = "Plantilla"
csv_path = "input/konus_catalog.csv"
output_path = "output/amazon_konus.xlsm"
start_row = 6
log_path = "logs/amazon_konus.log"
checkpoint_dir = Path("checkpoints")
checkpoint_dir.mkdir(parents=True, exist_ok=True)
checkpoint_file = checkpoint_dir / "amazon_konus.json"

load_dotenv()

MISTRAL_API_TOKEN = os.getenv("MISTRAL_API_TOKEN")
if not MISTRAL_API_TOKEN:
    raise RuntimeError("MISTRAL_API_TOKEN not found in environment")

# ---------------- LOGGING ----------------
Path(log_path).parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    filename=log_path,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logging.info("Script started")

# ---------------- LOAD FILES ----------------
Path(output_path).parent.mkdir(parents=True, exist_ok=True)
wb = load_workbook(excel_path, keep_vba=True)
ws = wb[sheet_name]

df = pd.read_csv(
    csv_path,
    sep=";",
    encoding="latin-1",
    engine="python"
)

# ---------------- AMAZON HEADERS ----------------
amazon_headers = [cell.value for cell in ws[4]]

# ---------------- CLEANERS ----------------
max_row = ws.max_row
if max_row >= start_row:
    ws.delete_rows(start_row, max_row - start_row + 1)


def clean_price(price_str):
    if not price_str:
        return None
    cleaned = re.sub(r"[^\d.]", "", str(price_str))
    return cleaned


def clean_json(text):
    match = re.search(r"\{.*\}", text, re.S)
    if not match:
        raise ValueError(f"Invalid JSON from LLM: {text}")
    return json.loads(match.group())


def direct_map(csv_row, enrichment):
    weight = csv_row.get("PesoNeto")
    if isinstance(weight, str):
        weight = weight.lower().replace("gr.", "").replace("gr", "").strip()

    medidas_raw = csv_row.get("Medidas")
    medidas = None
    if isinstance(medidas_raw, str):
        medidas_clean = medidas_raw.lower().replace("cm", "").replace(" ", "")
        medidas_parts = medidas_clean.split("x")
        medidas = "; ".join(medidas_parts)

    dims = enrichment.get("dimensions", {})

    return {
        "SKU": csv_row.get("EAN"),
        "SKU principal": csv_row.get("EAN"),
        "ID del producto": csv_row.get("EAN"),
        "Marca": csv_row.get("Marca"),
        "Nombre Modelo": csv_row.get("Modelo"),
        "Nombre del producto": csv_row.get("Título_producto"),
        "Palabra clave genérica": csv_row.get("Descripción_corta"),
        "Descripción del producto": csv_row.get("Descripción_larga"),
        "Peso Artículo": weight,
        "URL de la imagen principal": csv_row.get("Imagen_grande"),
        "Tipo de identificador del producto": "EAN",
        "Estado del producto": "Nuevo",
        "Precio de venta recomendado (PVPR)": clean_price(csv_row.get("PVP FINAL")),
        "Cumplimiento de código de canal (ES)": "DEFAULT",
        "Tu precio EUR (Vender en Amazon, ES)": clean_price(csv_row.get("PVP FINAL")),
        "Precio de venta. EUR (Vender en Amazon, ES)": clean_price(csv_row.get("PVP FINAL")),
        "Grupo de la marina mercante (ES)": "Nueva plantilla Envios",
        "Tamaño del anillo": medidas,
        "Número de Artículos": "1",
        "Numero de pieza": csv_row.get("Título_producto"),
        "Componentes Incluidos": "1 artículo",
        "Riesgo del GDPR": "No hay información electrónica almacenada.",
        "Tipo de producto": enrichment.get("product_type"),
        "Viñeta": enrichment.get("bullet"),
        "Grosor del artículo desde la parte delantera hasta la trasera": dims.get("thickness", {}).get("value"),
        "Unidad de altura del artículo": dims.get("height", {}).get("unit"),
        "Ancho del artículo de lado a lado": dims.get("width", {}).get("value"),
        "Unidad del ancho del artículo": dims.get("width", {}).get("unit"),
        "Aumento máximo": dims.get("max_magnification"),
        "Distancia focal mínima": dims.get("min_focal_distance", {}).get("value"),
        "Longitud Paquete": dims.get("package_length", {}).get("value"),
        "Unidad de longitud del paquete": dims.get("package_length", {}).get("unit"),
        "Ancho Paquete": dims.get("package_width", {}).get("value"),
        "Unidad de anchura del paquete": dims.get("package_width", {}).get("unit"),
        "Altura Paquete": dims.get("package_height", {}).get("value"),
        "Unidad de altura del paquete": dims.get("package_height", {}).get("unit"),
        "Peso del paquete": dims.get("package_weight", {}).get("value"),
        "Unidad del peso del paquete": dims.get("package_weight", {}).get("unit"),
        "Fabricante": "Konus",
        "¿Se necesitan baterías?": "No",
        "Normativas sobre mercancías peligrosas": "No aplicable",
        "Unidad de peso del artículo": "Gramos",
        "Número de cajas": "1"
    }

# ---------------- LOAD CHECKPOINT ----------------
if checkpoint_file.exists():
    with open(checkpoint_file, "r") as f:
        processed_skus = set(json.load(f))
else:
    processed_skus = set()

# ---------------- LLM ----------------
mistral = Mistral(api_key=MISTRAL_API_TOKEN)

ALLOWED_PRODUCT_TYPES = [
    "NAVIGATION_COMPASS",
    "FLASHLIGHT",
    "BINOCULAR",
    "TELESCOPE",
    "MAGNIFIER",
    "AIMING_SCOPE_SIGHT",
    "MICROSCOPES",
    "CAMERA_TRIPOD",
    "RANGEFINDER",
]

def classify_product_enrichment(csv_row, mistral):
    prompt = f"""
You are enriching Amazon product listings.

Rules:
- Choose EXACTLY ONE product type from the allowed list
- Generate EXACTLY ONE bullet point (short, factual, no marketing fluff)
- Infer dimensions ONLY if clearly implied; otherwise return null
- Units must be metric
- Return ONLY valid JSON
- No explanations, no markdown

Allowed product types:
{ALLOWED_PRODUCT_TYPES}

Product data:
- Title: {csv_row.get("Título_producto")}
- Short description: {csv_row.get("Descripción_corta")}
- Long description: {csv_row.get("Descripción_larga")}
- Family: {csv_row.get("Familia")}
- Model: {csv_row.get("Modelo")}

Return format:
{{
  "product_type": "ONE_OF_THE_ALLOWED_VALUES",
  "bullet": "One concise Amazon bullet point",
  "dimensions": {{
    "thickness": {{ "value": number|null, "unit": "cm"|null }},
    "height": {{ "value": number|null, "unit": "cm"|null }},
    "width": {{ "value": number|null, "unit": "cm"|null }},
    "max_magnification": number|null,
    "min_focal_distance": {{ "value": number|null, "unit": "cm"|null }},
    "package_length": {{ "value": number|null, "unit": "cm"|null }},
    "package_width": {{ "value": number|null, "unit": "cm"|null }},
    "package_height": {{ "value": number|null, "unit": "cm"|null }},
    "package_weight": {{ "value": number|null, "unit": "kg"|null }}
  }}
}}
"""

    res = mistral.chat.complete(
        model="mistral-small-latest",
        messages=[{"role": "user", "content": prompt}],
        stream=False,
    )

    data = clean_json(res.choices[0].message.content)

    if data.get("product_type") not in ALLOWED_PRODUCT_TYPES:
        raise ValueError(f"Invalid product type: {data.get('product_type')}")

    return data

# ---------------- PROCESS ROWS ----------------
current_row = start_row

for idx, csv_row in df.iterrows():
    csv_dict = csv_row.to_dict()
    sku = csv_dict.get("EAN")
    if sku in processed_skus:
        logging.info(f"Skipping already processed SKU {sku}")
        current_row += 1
        continue

    logging.info(f"Processing row {idx} SKU {sku}")

    try:
        enrichment = classify_product_enrichment(csv_dict, mistral)
    except Exception as e:
        logging.error(f"LLM enrichment failed for SKU {sku}: {e}")
        enrichment = {
            "product_type": None,
            "bullet": None,
            "dimensions": {}
        }

    mapped = direct_map(csv_dict, enrichment)

    for col_idx, header in enumerate(amazon_headers, start=1):
        value = mapped.get(header)
        if value is not None:
            ws.cell(row=current_row, column=col_idx, value=value)

    # Increment row for next entry
    current_row += 1

    # ---------------- SAVE CHECKPOINT ----------------
    processed_skus.add(sku)
    with open(checkpoint_file, "w") as f:
        json.dump(list(processed_skus), f)

    # ---------------- SAVE XLSM AFTER EACH ROW ----------------
    wb.save(output_path)

logging.info(f"Amazon XLSM generated: {output_path}")
print(f"Amazon XLSM generated: {output_path}")
